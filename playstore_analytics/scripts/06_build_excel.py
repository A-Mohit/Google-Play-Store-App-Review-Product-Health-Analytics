"""Stage 5 — Multi-app Excel validation workbook."""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "data" / "reviews_cleaned.csv"
OUT_PATH = ROOT / "exports" / "Music_Apps_Review_Analytics.xlsx"

def autosize(ws):
    for col in ws.columns:
        letter=get_column_letter(col[0].column)
        width=min(max(len(str(c.value or "")) for c in col)+2, 55)
        ws.column_dimensions[letter].width=width

def main():
    df=pd.read_csv(DATA_PATH, parse_dates=["review_date"])
    wb=Workbook(); raw=wb.active; raw.title="Raw_Reviews"
    cols=["app_id","app_name","review_id","user_name","rating","review_text","app_version","review_date","review_month","sentiment_score","sentiment_label","issue_category","thumbs_up","has_developer_reply","is_post_release_window"]
    raw.append(cols)
    for row in df[cols].itertuples(index=False, name=None): raw.append(list(row))
    raw.freeze_panes="A2"; raw.auto_filter.ref=raw.dimensions; autosize(raw)

    ws=wb.create_sheet("App_Summary")
    headers=["App Name","Reviews","Avg Rating","Avg Sentiment","Negative %","Positive %","Crash","Login","Payment","UI","Performance"]
    ws.append(headers)
    apps=sorted(df["app_name"].dropna().unique())
    for app in apps:
        g=df[df.app_name==app]
        ws.append([app,len(g),round(g.rating.mean(),2),round(g.sentiment_score.mean(),3),
                   round((g.sentiment_label=="negative").mean(),4),round((g.sentiment_label=="positive").mean(),4),
                   *[int((g.issue_category==x).sum()) for x in ["crash","login","payment","ui","performance"]]])
    for c in ws[1]: c.font=Font(bold=True); c.fill=PatternFill("solid", fgColor="1F4E78"); c.font=Font(bold=True,color="FFFFFF")
    for r in range(2,ws.max_row+1): ws.cell(r,5).number_format="0.0%"; ws.cell(r,6).number_format="0.0%"
    autosize(ws)

    dash=wb.create_sheet("KPI_Dashboard")
    dash["A1"]="Music Apps — Review Analytics Dashboard"; dash["A1"].font=Font(size=16,bold=True)
    dash.append([]); dash.append(["Metric","Value"]); dash.append(["Total Reviews",len(df)])
    dash.append(["Apps Compared",df.app_name.nunique()]); dash.append(["Overall Avg Rating",round(df.rating.mean(),2)])
    dash.append(["Overall Avg Sentiment",round(df.sentiment_score.mean(),3)])
    best=df.groupby("app_name").rating.mean().idxmax(); worst=df.groupby("app_name").rating.mean().idxmin()
    dash.append(["Highest Avg Rating App",best]); dash.append(["Lowest Avg Rating App",worst])
    dash.append(["Most Common Issue",df.loc[df.issue_category.ne("none"),"issue_category"].mode().iat[0] if (df.issue_category!="none").any() else "none"])
    chart=BarChart(); chart.title="Average Rating by App"; chart.y_axis.title="Rating"
    chart.add_data(Reference(ws,min_col=3,min_row=1,max_row=ws.max_row),titles_from_data=True)
    chart.set_categories(Reference(ws,min_col=1,min_row=2,max_row=ws.max_row)); chart.height=7; chart.width=12
    dash.add_chart(chart,"D3"); autosize(dash)

    wb.save(OUT_PATH); print(f"[excel] wrote {OUT_PATH} ({len(df)} rows, {df.app_name.nunique()} apps)")
if __name__ == "__main__": main()
